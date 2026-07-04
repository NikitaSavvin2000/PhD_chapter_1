import csv
import logging
from datetime import datetime
from zoneinfo import ZoneInfo
from pathlib import Path

from openpyxl import Workbook, load_workbook


def find_src_root(path: Path) -> Path:
    for parent in path.resolve().parents:
        if parent.name == "src":
            return parent
    return path.resolve()


SRC_ROOT = find_src_root(Path(__file__))


class MoscowFormatter(logging.Formatter):
    def format(self, record):
        dt = datetime.now(ZoneInfo("Europe/Moscow")).strftime("%d-%m-%Y %H:%M:%S")

        base = f"[{dt}] [{record.levelname}]"

        if record.levelno >= logging.WARNING:
            try:
                file_path = Path(record.pathname).resolve().relative_to(SRC_ROOT)
            except Exception:
                file_path = record.pathname

            base += f" [File: {file_path}]"

        return f"{base} {record.getMessage()}"


class ExcelLogHandler(logging.Handler):
    def __init__(self, log_dir: str):
        super().__init__()

        self.log_dir = Path(log_dir)
        self.log_dir.mkdir(parents=True, exist_ok=True)

        self.excel_path = self.log_dir / "logs.xlsx"

        if not self.excel_path.exists():
            workbook = Workbook()

            sheet = workbook.active
            sheet.title = "logs"

            sheet.append([
                "time",
                "level",
                "message"
            ])

            workbook.save(self.excel_path)

    def emit(self, record):
        dt = datetime.now(
            ZoneInfo("Europe/Moscow")
        ).strftime("%d-%m-%Y %H:%M:%S")

        try:
            workbook = load_workbook(self.excel_path)

            sheet = workbook["logs"]

            sheet.append([
                dt,
                record.levelname,
                record.getMessage()
            ])

            workbook.save(self.excel_path)

        except Exception as e:
            print(f"Excel logging error: {e}")


class CSVLogHandler(logging.Handler):
    def __init__(self, log_dir: str):
        super().__init__()

        self.log_dir = Path(log_dir)
        self.log_dir.mkdir(parents=True, exist_ok=True)

        self.csv_path = self.log_dir / "logs.csv"

        if not self.csv_path.exists():
            with open(self.csv_path, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow(["time", "level", "message"])

    def emit(self, record):
        try:
            dt = datetime.now(ZoneInfo("Europe/Moscow")).strftime("%d-%m-%Y %H:%M:%S")

            with open(self.csv_path, "a", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow([dt, record.levelname, record.getMessage()])

        except Exception as e:
            print(f"CSV logging error: {e}")


def get_logger(
        log_dir: str,
        name: str = "app"
) -> logging.Logger:

    logger = logging.getLogger(name)

    logger.setLevel(logging.INFO)
    logger.propagate = False

    if not logger.handlers:
        console_handler = logging.StreamHandler()

        console_handler.setLevel(logging.INFO)
        console_handler.setFormatter(MoscowFormatter())

        # excel_handler = ExcelLogHandler(log_dir)
        # excel_handler.setLevel(logging.INFO)

        csv_handler = CSVLogHandler(log_dir)
        csv_handler.setLevel(logging.INFO)

        logger.addHandler(console_handler)
        # logger.addHandler(excel_handler)
        logger.addHandler(csv_handler)

    return logger